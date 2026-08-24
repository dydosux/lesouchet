from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import io
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ocr_engine import extract_sheet_data
from wood_calc import calculate_sawn_volume, calculate_total_logs_volume

app = FastAPI(title="TimberOCR Local API", description="Local Offline Timber Sheet OCR & Salary Calculator")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/api/health")
def health():
    return {
        "status": "online",
        "engine": "OpenCV + Tesseract OCR Rus/Eng (100% Offline / Free / No Tokens)",
        "version": "1.1.0"
    }

@app.post("/api/recognize")
async def recognize_image(
    file: UploadFile = File(...),
    sawn_rate: float = Query(1600.0, description="Ставка за м3 готового пиломатериала (руб)"),
    slab_rate: float = Query(25.0, description="Ставка за 1 шт горбыля 2м (руб)")
):
    try:
        contents = await file.read()
        result = extract_sheet_data(contents, sawn_rate=sawn_rate, slab_rate=slab_rate)
        return JSONResponse(content=result)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")

class RecalculateRequest(BaseModel):
    shift_date: str
    shift_type: Optional[str] = "Дневная смена"
    brigade: Optional[str] = "Бригада №1"
    notes: Optional[str] = ""
    standard_table: List[Dict[str, Any]]
    extra_items: List[Dict[str, Any]]
    diameters: List[int]
    log_length: Optional[float] = 6.0
    sawn_rate_per_m3: Optional[float] = 1600.0
    slab_rate_per_piece: Optional[float] = 25.0

@app.post("/api/recalculate")
def recalculate(data: RecalculateRequest):
    try:
        total_g1_pcs = 0
        total_g2_pcs = 0
        total_g1_vol = 0.0
        total_g2_vol = 0.0
        
        updated_standard = []
        for row in data.standard_table:
            size = row.get("size", "")
            g1 = row.get("grade1_count")
            g2 = row.get("grade2_count")
            g1_c = int(g1) if g1 not in (None, "", "null") else 0
            g2_c = int(g2) if g2 not in (None, "", "null") else 0
            
            vol1 = calculate_sawn_volume(size, g1_c)
            vol2 = calculate_sawn_volume(size, g2_c)
            
            if g1_c > 0:
                total_g1_pcs += g1_c
                total_g1_vol += vol1
            if g2_c > 0:
                total_g2_pcs += g2_c
                total_g2_vol += vol2
                
            updated_standard.append({
                **row,
                "grade1_count": g1_c if g1_c > 0 else None,
                "grade2_count": g2_c if g2_c > 0 else None,
                "grade1_vol_m3": vol1,
                "grade2_vol_m3": vol2,
                "total_vol_m3": round(vol1 + vol2, 4),
                "is_filled": (g1_c > 0 or g2_c > 0)
            })
            
        total_extra_pcs = 0
        total_extra_vol = 0.0
        sawn_extra_vol = 0.0
        total_slab_pcs = 0
        updated_extra = []
        
        for extra in data.extra_items:
            cnt = int(extra.get("count", 0) or 0)
            sz = extra.get("size", "")
            is_slab = extra.get("is_slab", False)
            vol = 0.0
            if is_slab:
                vol = round(cnt * 0.015, 3)
                total_slab_pcs += cnt
            else:
                vol = calculate_sawn_volume(sz, cnt)
                sawn_extra_vol += vol
                
            total_extra_pcs += cnt
            total_extra_vol += vol
            updated_extra.append({
                **extra,
                "count": cnt,
                "vol_m3": vol
            })
            
        logs_res = calculate_total_logs_volume(data.diameters, data.log_length or 6.0)
        
        sawn_base_vol = round(total_g1_vol + total_g2_vol + sawn_extra_vol, 3)
        total_sawn_vol = round(total_g1_vol + total_g2_vol + total_extra_vol, 3)
        total_logs_vol = logs_res["total_volume_m3"]
        
        raw_yield = (sawn_base_vol / total_logs_vol * 100) if total_logs_vol > 0 else 0.0
        yield_floor = math.floor(raw_yield)
        
        sawn_rate = data.sawn_rate_per_m3 or 1600.0
        slab_rate = data.slab_rate_per_piece or 25.0
        sawn_salary = int(round(sawn_base_vol * sawn_rate))
        slab_salary = int(round(total_slab_pcs * slab_rate))
        total_salary = sawn_salary + slab_salary
        
        return {
            "metadata": {
                "shift_date": data.shift_date,
                "shift_type": data.shift_type or "Дневная смена",
                "brigade": data.brigade or "Бригада №1",
                "notes": data.notes or "",
                "model_engine": "Local Calculator Engine"
            },
            "summary": {
                "total_grade1_count": total_g1_pcs,
                "total_grade2_count": total_g2_pcs,
                "total_pieces": total_g1_pcs + total_g2_pcs + total_extra_pcs,
                "total_grade1_volume_m3": round(total_g1_vol, 3),
                "total_grade2_volume_m3": round(total_g2_vol, 3),
                "total_extra_volume_m3": round(total_extra_vol, 3),
                "total_sawn_base_volume_m3": sawn_base_vol,
                "total_sawn_volume_m3": total_sawn_vol,
                "total_logs_count": logs_res["count"],
                "total_logs_volume_m3": total_logs_vol,
                "raw_yield_percent": round(raw_yield, 1),
                "yield_percent": yield_floor,
                "salary": {
                    "sawn_rate_per_m3": sawn_rate,
                    "slab_rate_per_piece": slab_rate,
                    "sawn_base_volume_m3": sawn_base_vol,
                    "sawn_salary_rub": sawn_salary,
                    "slab_count": total_slab_pcs,
                    "slab_salary_rub": slab_salary,
                    "total_salary_rub": total_salary
                }
            },
            "standard_table": updated_standard,
            "extra_items": updated_extra,
            "roundwood_logs": {
                "diameters": data.diameters,
                "count": logs_res["count"],
                "total_volume_m3": logs_res["total_volume_m3"],
                "breakdown": logs_res["breakdown"]
            }
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/export/excel")
def export_excel(data: RecalculateRequest):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводка и Зарплата"
    
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sub_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    bold_font = Font(name="Arial", size=11, bold=True)
    regular_font = Font(name="Arial", size=11)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells("A1:G1")
    ws["A1"] = f"РАСЧЁТНАЯ ВЕДОМОСТЬ РАСПИЛОВКИ И ЗАРПЛАТЫ — СМЕНА {data.shift_date} ({data.brigade or 'Бригада'})"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = center_align
    
    ws.append([])
    ws.append(["Размер (мм)", "1 СОРТ (шт)", "Объем 1с (м³)", "2 СОРТ (шт)", "Объем 2с (м³)", "Всего шт", "Итого объем (м³)"])
    for col_idx in range(1, 8):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = sub_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    total_pcs_g1 = 0
    total_pcs_g2 = 0
    total_vol_g1 = 0.0
    total_vol_g2 = 0.0
    sawn_extra_vol = 0.0
    slab_count = 0
    
    current_row = 4
    for row in data.standard_table:
        g1 = int(row.get("grade1_count") or 0)
        g2 = int(row.get("grade2_count") or 0)
        if g1 > 0 or g2 > 0:
            sz = row.get("size", "")
            vol1 = calculate_sawn_volume(sz, g1)
            vol2 = calculate_sawn_volume(sz, g2)
            total_pcs_g1 += g1
            total_pcs_g2 += g2
            total_vol_g1 += vol1
            total_vol_g2 += vol2
            
            ws.append([sz, g1 if g1 > 0 else "-", round(vol1, 4) if vol1 > 0 else "-", g2 if g2 > 0 else "-", round(vol2, 4) if vol2 > 0 else "-", g1 + g2, round(vol1 + vol2, 4)])
            for col_idx in range(1, 8):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = regular_font
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1

    for extra in data.extra_items:
        cnt = int(extra.get("count") or 0)
        if cnt > 0:
            sz = extra.get("size", "")
            name = extra.get("name", sz)
            is_slab = extra.get("is_slab", False)
            if is_slab:
                vol = round(cnt * 0.015, 3)
                slab_count += cnt
            else:
                vol = calculate_sawn_volume(sz, cnt)
                sawn_extra_vol += vol
                
            total_pcs_g1 += cnt
            total_vol_g1 += vol
            ws.append([f"{name} (доп.)", cnt, round(vol, 4), "-", "-", cnt, round(vol, 4)])
            for col_idx in range(1, 8):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = regular_font
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1
            
    ws.append(["ИТОГО ПРОДУКЦИЯ:", total_pcs_g1, round(total_vol_g1, 3), total_pcs_g2, round(total_vol_g2, 3), total_pcs_g1 + total_pcs_g2, round(total_vol_g1 + total_vol_g2, 3)])
    for col_idx in range(1, 8):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = bold_font
        cell.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
        cell.alignment = center_align
        cell.border = thin_border
    current_row += 2
    
    # Salary block
    sawn_base = round(total_vol_g1 + total_vol_g2 - (slab_count * 0.015), 3) if slab_count > 0 else round(total_vol_g1 + total_vol_g2, 3)
    sawn_rate = data.sawn_rate_per_m3 or 1600.0
    slab_rate = data.slab_rate_per_piece or 25.0
    sawn_sal = int(round(sawn_base * sawn_rate))
    slab_sal = int(round(slab_count * slab_rate))
    tot_sal = sawn_sal + slab_sal
    
    ws.cell(row=current_row, column=1, value="РАСЧЁТ ЗАРПЛАТЫ БРИГАДЫ:").font = bold_font
    current_row += 1
    
    ws.append(["Пиломатериал готовый:", f"{sawn_base} м³", f"× {sawn_rate} руб/м³", f"= {sawn_sal} руб", "", "", ""])
    current_row += 1
    ws.append(["Горбыль 2м (штакет/доска):", f"{slab_count} шт", f"× {slab_rate} руб/шт", f"= {slab_sal} руб", "", "", ""])
    current_row += 1
    
    ws.append(["ИТОГО К ВЫДАЧЕ БРИГАДЕ:", "", "", f"{tot_sal} РУБЛЕЙ", "", "", ""])
    for col_idx in range(1, 5):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = Font(name="Arial", size=12, bold=True, color="166534")
        cell.fill = PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid")
        cell.alignment = center_align
        cell.border = thin_border
    current_row += 2
    
    # Raw logs
    logs_res = calculate_total_logs_volume(data.diameters, data.log_length or 6.0)
    raw_yield = (sawn_base / logs_res["total_volume_m3"] * 100) if logs_res["total_volume_m3"] > 0 else 0
    yield_floor = math.floor(raw_yield)
    
    ws.cell(row=current_row, column=1, value=f"ИСХОДНЫЙ КРУГЛЯК: {logs_res['count']} шт ({logs_res['total_volume_m3']} м³). ВЫХОД ЛЕСА: {yield_floor}% (округлено в меньшую сторону)").font = bold_font
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=timber_salary_{data.shift_date}.xlsx"}
    )
